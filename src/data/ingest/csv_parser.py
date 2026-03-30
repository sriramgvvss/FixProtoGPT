"""
Module: src.data.ingest.csv_parser
====================================

Extract FIX protocol tag/field definitions from CSV/TSV spreadsheets.

Typical Columns
---------------
* Tag (or Number) — integer field tag
* Name — human-readable field name
* Type (or Data Type) — FIX data type
* Required (or Req'd) — Y/N
* Description — free-text comment

Also supports ``.xlsx`` files via ``openpyxl`` (optional).

Coding Standards
----------------
- PEP 8  : Python Style Guide
- PEP 257 : Docstring Conventions — Google-style docstrings
- PEP 484 : Type Hints
- Google Python Style Guide

Author : FixProtoGPT Team
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Dict, List, Optional

from src.data.ingest.base import (
    CanonicalSpec,
    SpecKind,
    SpecParser,
    register_parser,
)

logger = logging.getLogger(__name__)


@register_parser
class CSVParser(SpecParser):
    """Extract FIX field definitions from CSV, TSV, and XLSX files."""

    CSV_EXTENSIONS = {".csv", ".tsv", ".txt"}
    XLSX_EXTENSIONS = {".xlsx", ".xls"}

    def can_handle(self, path: Path) -> bool:
        """Return ``True`` for CSV/TSV/XLSX files."""
        return path.suffix.lower() in (self.CSV_EXTENSIONS | self.XLSX_EXTENSIONS)

    def parse(self, path: Path) -> List[CanonicalSpec]:
        """Parse a CSV/TSV/XLSX and return field definitions.

        Args:
            path: Path to the file.

        Returns:
            List of :class:`CanonicalSpec` records.
        """
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        suffix = path.suffix.lower()
        if suffix in self.XLSX_EXTENSIONS:
            return self._parse_xlsx(path)
        return self._parse_csv(path)

    # ── CSV / TSV ─────────────────────────────────────────────────

    def _parse_csv(self, path: Path) -> List[CanonicalSpec]:
        """Parse a plain CSV or TSV file."""
        logger.info("Parsing CSV spec: %s", path.name)

        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        specs: List[CanonicalSpec] = []

        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            header_row = next(reader, None)
            if header_row is None:
                return []

            header = [h.strip().lower() for h in header_row]
            col_map = self._map_columns(header)
            if "tag" not in col_map:
                logger.warning("No 'tag' column found in %s", path.name)
                return []

            for row_num, row in enumerate(reader, start=2):
                spec = self._row_to_spec(row, col_map, path.name, row_num)
                if spec is not None:
                    specs.append(spec)

        logger.info("CSV %s → %d records", path.name, len(specs))
        return specs

    # ── XLSX ──────────────────────────────────────────────────────

    def _parse_xlsx(self, path: Path) -> List[CanonicalSpec]:
        """Parse an Excel spreadsheet using ``openpyxl``."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX parsing. "
                "Install it with: pip install openpyxl"
            ) from exc

        logger.info("Parsing XLSX spec: %s", path.name)
        specs: List[CanonicalSpec] = []

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = [str(c).strip().lower() if c else "" for c in rows[0]]
            col_map = self._map_columns(header)
            if "tag" not in col_map:
                continue

            for row_num, row in enumerate(rows[1:], start=2):
                cells = [str(c).strip() if c else "" for c in row]
                spec = self._row_to_spec(cells, col_map, path.name, row_num)
                if spec is not None:
                    specs.append(spec)

        wb.close()
        logger.info("XLSX %s → %d records", path.name, len(specs))
        return specs

    # ── Shared helpers ────────────────────────────────────────────

    def _row_to_spec(
        self,
        row: list,
        col_map: Dict[str, int],
        source: str,
        row_num: int,
    ) -> Optional[CanonicalSpec]:
        """Convert a spreadsheet row into a :class:`CanonicalSpec`."""
        tag_str = self._cell(row, col_map.get("tag"))
        tag = self._safe_int(tag_str)
        if tag is None:
            return None

        name = self._cell(row, col_map.get("name"))
        dtype = self._cell(row, col_map.get("type"))
        req = self._cell(row, col_map.get("required"))
        desc = self._cell(row, col_map.get("description"))
        values_str = self._cell(row, col_map.get("values"))

        # Parse comma-separated enum values like "1=Buy,2=Sell"
        values: Dict[str, str] = {}
        if values_str:
            for pair in values_str.split(","):
                pair = pair.strip()
                if "=" in pair:
                    k, v = pair.split("=", 1)
                    values[k.strip()] = v.strip()

        return CanonicalSpec(
            kind=SpecKind.FIELD,
            tag=tag,
            name=name,
            data_type=dtype.upper() if dtype else None,
            required=self._parse_bool(req),
            description=desc,
            values=values,
            source=f"csv:{source}",
            meta={"row": row_num},
        )

    @staticmethod
    def _map_columns(header: List[str]) -> Dict[str, int]:
        """Map CSV header names to logical field roles (tag, name, type, etc.)."""
        mapping: Dict[str, int] = {}
        for idx, col in enumerate(header):
            if "tag" in col or "number" in col:
                mapping.setdefault("tag", idx)
            elif "name" in col or "field" in col:
                mapping.setdefault("name", idx)
            elif "type" in col:
                mapping.setdefault("type", idx)
            elif "req" in col:
                mapping.setdefault("required", idx)
            elif "desc" in col or "comment" in col:
                mapping.setdefault("description", idx)
            elif "value" in col or "enum" in col:
                mapping.setdefault("values", idx)
        return mapping

    @staticmethod
    def _cell(row: list, idx: Optional[int]) -> str:
        """Safely extract a stripped string from a row by column index."""
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] else ""

    @staticmethod
    def _safe_int(s: str) -> Optional[int]:
        """Parse a string to int, tolerating Excel-style floats like '55.0'."""
        try:
            return int(float(s))  # handles "55.0" from Excel
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_bool(s: str) -> Optional[bool]:
        """Interpret common truthy strings (y, yes, required, true, 1) as booleans."""
        if not s:
            return None
        return s.strip().lower() in {"y", "yes", "required", "true", "1"}
